#!/usr/bin/python
# -*- #################
#Author:Shariful Islam
#Contact: msi_g@yahoo.com

'''
This tool reads an excel file that contains the scaling factor for a raster and 
finds the appropriate raster to multiply with the proper scaling factor.
e.g.
In the input solar radiation raster folder there are 154 esri grid raster that represent the 154 days of any year and the corresponding 
excel scaling factor file contains the scale factor for those 154 days for multiple year. So now using this tool you can create solar radiation raster for
all the years in the excel file and this raster for each 154 days.

Say raster name in 0512 i.e. May 12 say again excel file has two years 1912 and 1913
So now you can generate two raster for May 12 for both 1912 and 1913.

'''

import arcpy,os,sys,datetime,re,shutil
from arcpy.sa import Con
from arcpy import Raster
from arcpy.sa import Times
from openpyxl import load_workbook
arcpy.env.overwriteOutput = True
arcpy.CheckOutExtension("spatial")

use = 'T' #'S'

INPUT_RASTER_FOLDER = (arcpy.GetParameterAsText(0) if use =='T' else
                     r"C:\Users\Winrock\Desktop\Ryan\RasterScaling\Check02Sept\SolarRadiationRaster\Rasters"
                     )
INPUT_EXCEL_SCALING_FILE = (arcpy.GetParameterAsText(1)  if use =='T' else
                     #r"C:\Users\Winrock\Desktop\Ryan\RasterScaling\Check02Sept\Solar radiation scaling factor 2007-2012.xlsx"
                     r"C:\Users\Winrock\Desktop\Ryan\RasterScaling\Check02Sept\Solar radiation scaling factor average 1912-2006.xlsx"
                     )
OUTPUT_FOLDER = (arcpy.GetParameterAsText(2)  if use =='T' else
                     r"C:\Users\Winrock\Desktop\Ryan\RasterScaling\Check02Sept\SolarRadiationRaster\Output"
                     )

TEMP_FOLDER_PATH = (arcpy.GetParameterAsText(3)  if use =='T' else
                    r"C:\Users\Winrock\Desktop\Ryan\RasterScaling\Check02Sept\mytemp"
                    )



temp_db_pattern = r'Scratch_.+'
Scaling_Data=[]
solr_rad_files = [p for p in os.listdir(INPUT_RASTER_FOLDER)]# if p.endswith('.tif')]

snow_wb = load_workbook(filename=INPUT_EXCEL_SCALING_FILE, read_only=True)
snow_ws = snow_wb[snow_wb.sheetnames[0]]
    
for row in snow_ws.rows:
    d = []
    if len(row)>1:
        for cell in row:
            if cell.value == None:
                pass
            elif cell.value == 0:
                d.append(0.000000)
            else:
                d.append(cell.value)
        Scaling_Data.append(d)
Scaling_Data = Scaling_Data[1:]
seen = set()
Scaling_Data = [x for x in Scaling_Data if x[0] not in seen and not seen.add(x[0])]# removing duplicate date
Scaling_Data = [[str(roww[0].strftime("%Y%m%d")).strip(), roww[1]] for roww in Scaling_Data]

available_years = list(set([y[0][0:4] for y in Scaling_Data]))
if use =='T':
    arcpy.AddMessage("\n\nOnly found years  %s"%', '.join(available_years))
else:
    print ("\n\nOnly found years  %s"%', '.join(available_years))
#folder deleter
def purge(dirpth, pattern):
    for f in os.listdir(dirpth):
        if re.search(pattern, f):
            pth = os.path.join(dirpth, f)
            shutil.rmtree(pth, ignore_errors=True)
 
if use =='T':
    arcpy.AddMessage('\nDoing the scaling........\n')
else:
    print ('\nDoing the scaling........\n')
for rastr in solr_rad_files:
    scratch_db_name = "Scratch_"+str(rastr)
    arcpy.CreateFileGDB_management(out_folder_path=TEMP_FOLDER_PATH, out_name=scratch_db_name, out_version="CURRENT")
    scr_db = os.path.join(TEMP_FOLDER_PATH,scratch_db_name+".gdb")
    arcpy.env.scratchWorkspace = scr_db    
    for yearitem in Scaling_Data:
        folder_name = yearitem[0][0:4]
        image_name = yearitem[0][-4:]
        sclfactor = yearitem[1]
        if image_name == rastr:
            input_raster_path = os.path.join(INPUT_RASTER_FOLDER,rastr)
            input_raster = Raster(input_raster_path)
            out_folder_path = os.path.join(OUTPUT_FOLDER,folder_name)
            if not os.path.exists(out_folder_path):
                os.mkdir(out_folder_path)
            OutputRasterPath = os.path.join(OUTPUT_FOLDER,folder_name,'g'+image_name)
            outpt = Times(input_raster, sclfactor)
            outpt.save(OutputRasterPath)

purge(TEMP_FOLDER_PATH, temp_db_pattern)
if use =='T':
    arcpy.AddMessage('\nAll Done!')
else:
    print('\nAll Done!')